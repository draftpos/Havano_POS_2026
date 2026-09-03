import os
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QMessageBox, QFileDialog, QProgressDialog, QApplication, QFrame, QWidget
)
from PySide6.QtCore import Qt, QStandardPaths, QThread, Signal

# Patch numpy 2.0+ compatibility for openpyxl
try:
    import numpy
    for _k, _v in [('short', getattr(numpy, 'int16', int)), ('ushort', getattr(numpy, 'uint16', int)),
                   ('intc', getattr(numpy, 'int32', int)), ('uintc', getattr(numpy, 'uint32', int)),
                   ('int_', getattr(numpy, 'int64', int)), ('uint', getattr(numpy, 'uint64', int)),
                   ('half', getattr(numpy, 'float16', float)), ('single', getattr(numpy, 'float32', float)),
                   ('double', getattr(numpy, 'float64', float)), ('longdouble', getattr(numpy, 'float64', float))]:
        if not hasattr(numpy, _k):
            setattr(numpy, _k, _v)
except Exception:
    pass

try:
    import openpyxl
except ImportError:
    openpyxl = None

from theme import NAVY, WHITE, ACCENT, ACCENT_H, SUCCESS, SUCCESS_H, BORDER, MUTED
from models.product import (
    get_product_by_part_no, update_product, create_product,
    upsert_item_price, get_products_by_part_nos
)


class StockImportWorker(QThread):
    progress = Signal(int, str)          # (rows processed, current part no preview)
    finished_ok = Signal(int, list)      # (success count, error list)
    failed = Signal(str)                 # unrecoverable error

    def __init__(self, rows, columns, parent=None):
        super().__init__(parent)
        self.rows = rows          # data rows (excludes header)
        self.columns = columns    # dict of resolved column indices
        self._canceled = False

    def cancel(self):
        self._canceled = True

    def run(self):
        cols = self.columns
        part_col = cols['part_col']
        stock_col = cols['stock_col']

        success = 0
        errors = []
        total_rows = len(self.rows)

        all_part_nos = []
        for row in self.rows:
            if part_col is not None and len(row) > part_col and row[part_col] is not None:
                pn = str(row[part_col]).strip().upper()
                if pn and pn != 'NAN':
                    all_part_nos.append(pn)

        existing_lookup = {}
        try:
            existing_lookup = get_products_by_part_nos(all_part_nos)
        except Exception:
            existing_lookup = None

        for idx, row in enumerate(self.rows):
            if self._canceled:
                break

            if idx % 5 == 0 or idx == total_rows - 1:
                part_preview = (
                    str(row[part_col]).strip().upper()
                    if part_col is not None and len(row) > part_col and row[part_col] else ""
                )
                self.progress.emit(idx + 1, part_preview)

            try:
                part_val_raw = row[part_col]
                if part_val_raw is None:
                    continue
                part_no = str(part_val_raw).strip().upper()
                if not part_no or part_no == 'NAN':
                    continue

                def get_val(col, default, cast_func):
                    if col is not None and len(row) > col and row[col] is not None:
                        try:
                            return cast_func(row[col])
                        except Exception:
                            return default
                    return default

                def get_str(col, default=""):
                    if col is not None and len(row) > col and row[col] is not None:
                        return str(row[col]).strip()
                    return default

                def get_bool(col, default):
                    if col is not None and len(row) > col and row[col] is not None:
                        val = str(row[col]).strip().lower()
                        return val in ['1', 'true', 'yes', 'y', 't']
                    return default

                stock_val = get_val(cols['stock_col'], 0.0, float)
                price_val = get_val(cols['price_col'], 0.0, float)
                cost_val = get_val(cols['cost_col'], 0.0, float)
                conv_val = get_val(cols['conv_col'], 1.0, float)
                tax_rate_val = get_val(cols['tax_rate_col'], 0.0, float)
                reorder_val = get_val(cols['reorder_col'], 0.0, float)

                name_val = get_str(cols['name_col'], part_no)
                desc_val = get_str(cols['desc_col'], "")
                cat_val = get_str(cols['cat_col'], "General")
                uom_val = get_str(cols['uom_col'], "Unit")
                tax_type_val = get_str(cols['tax_type_col'], "VAT")
                batch_val = get_str(cols['batch_col'], "")
                expiry_val = get_str(cols['expiry_col'], "")

                pharmacy_val = get_bool(cols['pharmacy_col'], False)
                butchery_val = get_bool(cols['butchery_col'], False)
                track_val = get_bool(cols['track_col'], True)
                bundle_val = get_bool(cols['bundle_col'], False)

                if existing_lookup is not None:
                    prod = existing_lookup.get(part_no)
                else:
                    prod = get_product_by_part_no(part_no)

                if prod:
                    update_product(
                        prod['id'],
                        stock=stock_val,
                        price=price_val,
                        name=name_val if cols['name_col'] else None,
                        description=desc_val if cols['desc_col'] else None,
                        category=cat_val if cols['cat_col'] else None,
                        uom=uom_val if cols['uom_col'] else None,
                        conversion_factor=conv_val if cols['conv_col'] else None,
                        cost_price=cost_val if cols['cost_col'] else None,
                        tax_type=tax_type_val if cols['tax_type_col'] else None,
                        tax_rate=tax_rate_val if cols['tax_rate_col'] else None,
                        reorder_level=reorder_val if cols['reorder_col'] else None,
                        is_pharmacy_product=pharmacy_val if cols['pharmacy_col'] else None,
                        is_butchery_product=butchery_val if cols['butchery_col'] else None,
                        track_stock=track_val if cols['track_col'] else None,
                        is_product_bundle=bundle_val if cols['bundle_col'] else None,
                        batch_no=batch_val if cols['batch_col'] else "",
                        expiry_date=expiry_val if cols['expiry_col'] else ""
                    )
                    upsert_item_price(prod['part_no'], "Standard Selling", prod.get('uom', uom_val), price_val)
                else:
                    p = create_product(
                        part_no=part_no,
                        name=name_val,
                        price=price_val,
                        stock=stock_val,
                        category=cat_val,
                        uom=uom_val,
                        conversion_factor=conv_val,
                        cost_price=cost_val,
                        description=desc_val,
                        tax_type=tax_type_val,
                        tax_rate=tax_rate_val,
                        reorder_level=reorder_val,
                        is_pharmacy_product=pharmacy_val,
                        is_butchery_product=butchery_val,
                        track_stock=track_val,
                        is_product_bundle=bundle_val,
                        batch_no=batch_val,
                        expiry_date=expiry_val
                    )
                    if p:
                        upsert_item_price(part_no, "Standard Selling", uom_val, price_val)

                success += 1
            except Exception as e:
                errors.append(f"Row {idx + 2}: {e}")

        self.finished_ok.emit(success, errors)


class UploadStockDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Bulk Stock Upload")
        self.setFixedSize(460, 240)
        self.setStyleSheet(f"QDialog {{ background: {WHITE}; }}")
        self._worker = None
        self._loader = None
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(14)

        title = QLabel("Bulk Product & Stock Upload")
        title.setStyleSheet(f"color: {NAVY}; font-size: 18px; font-weight: bold; background: transparent;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        desc = QLabel(
            "Download the template to add or update your products.\n"
            "Fill out the columns you need. 'Part No' and 'Stock' are required."
        )
        desc.setStyleSheet("color: #475569; font-size: 13px; background: transparent;")
        desc.setAlignment(Qt.AlignCenter)
        layout.addWidget(desc)

        layout.addStretch()

        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(14)

        self.btn_download = QPushButton("Download Template")
        self.btn_download.setFixedHeight(38)
        self.btn_download.setCursor(Qt.PointingHandCursor)
        self.btn_download.setStyleSheet(f"""
            QPushButton {{
                background-color: {ACCENT}; color: white; font-weight: bold;
                border-radius: 5px; font-size: 12px; padding: 0 12px;
            }}
            QPushButton:hover {{ background-color: {ACCENT_H}; }}
        """)
        self.btn_download.clicked.connect(self._download_template)
        btn_layout.addWidget(self.btn_download)

        self.btn_upload = QPushButton("Upload Excel File")
        self.btn_upload.setFixedHeight(38)
        self.btn_upload.setCursor(Qt.PointingHandCursor)
        self.btn_upload.setStyleSheet(f"""
            QPushButton {{
                background-color: {SUCCESS}; color: white; font-weight: bold;
                border-radius: 5px; font-size: 12px; padding: 0 12px;
            }}
            QPushButton:hover {{ background-color: {SUCCESS_H}; }}
            QPushButton:disabled {{ background-color: #cbd5e1; color: #94a3b8; }}
        """)
        self.btn_upload.clicked.connect(self._upload_file)
        btn_layout.addWidget(self.btn_upload)

        self.btn_close = QPushButton("Close")
        self.btn_close.setFixedHeight(38)
        self.btn_close.setCursor(Qt.PointingHandCursor)
        self.btn_close.setStyleSheet("""
            QPushButton {
                background-color: #f1f5f9; color: #475569; font-weight: bold;
                border: 1px solid #cbd5e1; border-radius: 5px; font-size: 12px; padding: 0 12px;
            }
            QPushButton:hover { background-color: #e2e8f0; color: #1e293b; }
        """)
        self.btn_close.clicked.connect(self.reject)
        btn_layout.addWidget(self.btn_close)

        layout.addLayout(btn_layout)

    def _download_template(self):
        if openpyxl is None:
            QMessageBox.critical(self, "Dependencies Missing", "The 'openpyxl' library is required to create Excel files.")
            return

        docs = QStandardPaths.writableLocation(QStandardPaths.DocumentsLocation)
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Template", os.path.join(docs, "Stock_Upload_Template.xlsx"), "Excel Files (*.xlsx)"
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Products"
            headers = [
                "Part No", "Product Name", "Description", "Category", "Stock",
                "UOM", "Conversion Factor", "Cost Price", "Price", "Tax Type",
                "Tax Rate", "Reorder Level", "Is Pharmacy Product",
                "Is Butchery Product", "Track Stock", "Is Product Bundle",
                "Batch No", "Expiry Date"
            ]
            ws.append(headers)
            ws.append(["ITEM-001", "Example Product A", "A sample product", "General", 10.0, "Unit", 1.0, 10.0, 15.0, "VAT", 15.0, 5.0, 0, 0, 1, 0, "", ""])
            ws.append(["ITEM-002", "Example Product B", "", "General", 50.0, "Unit", 1.0, 18.0, 25.0, "ZERO RATED", 0.0, 10.0, 0, 0, 1, 0, "", ""])
            wb.save(path)
            QMessageBox.information(self, "Success", f"Template saved successfully to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save template:\n{e}")

    def _upload_file(self):
        if openpyxl is None:
            QMessageBox.critical(self, "Dependencies Missing", "The 'openpyxl' library is required to read Excel files.")
            return

        file_path, _ = QFileDialog.getOpenFileName(
            self, "Import Stock from Excel", "", "Excel Files (*.xlsx);;All Files (*)"
        )
        if not file_path:
            return

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to read Excel file:\n{e}")
            return

        if not rows or len(rows) < 2:
            QMessageBox.critical(self, "Import Error", "Excel file is empty or missing data.")
            return

        header = [str(c).lower().strip() if c is not None else "" for c in rows[0]]
        df_columns = {k: i for i, k in enumerate(header) if k}

        def get_col(*keywords):
            for k in keywords:
                for df_k, idx in df_columns.items():
                    if df_k == k:
                        return idx
            for k in keywords:
                for df_k, idx in df_columns.items():
                    if k in df_k:
                        return idx
            return None

        columns = {
            'part_col': get_col("part no", "part", "code"),
            'name_col': get_col("product name", "name"),
            'desc_col': get_col("description", "desc"),
            'cat_col': get_col("category", "cat"),
            'stock_col': get_col("stock", "qty", "quantity"),
            'uom_col': get_col("uom", "unit"),
            'conv_col': get_col("conversion"),
            'cost_col': get_col("cost"),
            'tax_type_col': get_col("tax type"),
            'tax_rate_col': get_col("tax rate"),
            'reorder_col': get_col("reorder"),
            'pharmacy_col': get_col("pharmacy"),
            'butchery_col': get_col("butchery"),
            'track_col': get_col("track"),
            'bundle_col': get_col("bundle"),
            'batch_col': get_col("batch"),
            'expiry_col': get_col("expiry"),
        }

        price_col = None
        for k, idx in df_columns.items():
            if "price" in k and "cost" not in k:
                price_col = idx
                break
        if price_col is None:
            price_col = get_col("price")
        columns['price_col'] = price_col

        if columns['part_col'] is None or columns['stock_col'] is None:
            QMessageBox.critical(self, "Import Error", "Excel must contain at least 'Part No' and 'Stock' columns.")
            return

        total_rows = len(rows) - 1
        from views.components.smart_progress_dialog import SmartProgressDialog
        self._loader = SmartProgressDialog(title="Bulk Stock Excel Import", total_items=total_rows, parent=self)
        self._loader.show()

        self.btn_upload.setEnabled(False)

        self._worker = StockImportWorker(rows[1:], columns, parent=self)
        self._worker.progress.connect(self._on_progress)
        self._worker.finished_ok.connect(self._on_finished)

        if hasattr(self._loader, 'canceled'):
            self._loader.canceled.connect(self._worker.cancel)

        self._worker.start()

    def _on_progress(self, count, part_preview):
        if self._loader:
            if getattr(self._loader, "cancelled", False):
                if self._worker:
                    self._worker.cancel()
            else:
                self._loader.update_progress(count, part_preview)

    def _on_finished(self, success, errors):
        was_canceled = False
        if self._loader:
            if getattr(self._loader, "cancelled", False):
                was_canceled = True
            self._loader.accept()
        if self._worker and getattr(self._worker, "_canceled", False):
            was_canceled = True

        self.btn_upload.setEnabled(True)

        if was_canceled:
            QMessageBox.warning(self, "Import Cancelled", f"The Excel import was cancelled after processing {success} items.")
        else:
            msg = f"Successfully processed {success} items."
            if errors:
                msg += f"\nErrors ({len(errors)}):\n" + "\n".join(errors[:10])
                if len(errors) > 10:
                    msg += "\n..."
            QMessageBox.information(self, "Import Completed", msg)
            self.accept()